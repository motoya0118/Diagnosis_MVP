from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any, Iterable, Sequence
from sqlalchemy import delete, inspect, select
from sqlalchemy.orm import Session

if TYPE_CHECKING:  # pragma: no cover - import for type checking only
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
else:  # pragma: no cover - fallbacks when openpyxl is unavailable at runtime
    Workbook = Any  # type: ignore[assignment]
    Worksheet = Any  # type: ignore[assignment]

from app.core.errors import ErrorCode
from app.core.registry import resolve_outcome_model
from app.models.diagnostic import (
    Diagnostic,
    DiagnosticVersion,
    Option,
    Question,
    VersionOption,
    VersionOutcome,
    VersionQuestion,
)
from app.services.diagnostics.audit import record_diagnostic_version_log
from app.services.diagnostics.template_exporter import TemplateExporter


REQUIRED_SHEETS = ("questions", "options", "outcomes")


def _load_workbook(stream: io.BytesIO) -> Any:
    from openpyxl import load_workbook as _load_workbook

    return _load_workbook(stream, data_only=True)


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter as _get_column_letter

    return _get_column_letter(index)


@dataclass(frozen=True)
class QuestionImportRow:
    q_code: str
    display_text: str
    multi: bool
    sort_order: int
    is_active: bool
    row_index: int = 0


@dataclass(frozen=True)
class OptionImportRow:
    q_code: str
    opt_code: str
    display_label: str
    sort_order: int
    is_active: bool
    llm_op: dict[str, Any] | None = None
    row_index: int = 0


@dataclass(frozen=True)
class OutcomeImportRow:
    values: dict[str, Any]
    row_index: int = 0


@dataclass(frozen=True)
class StructureImportBatch:
    questions: list[QuestionImportRow]
    options: list[OptionImportRow]
    outcomes: list[OutcomeImportRow]
    warnings: list[str] = field(default_factory=list)
    outcome_headers: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class StructureImportSummary:
    version_id: int
    questions_imported: int
    options_imported: int
    outcomes_imported: int
    warnings: list[str]


class StructureImportParseError(Exception):
    def __init__(
        self,
        *,
        error_code: ErrorCode,
        detail: str | None = None,
        invalid_cells: Sequence[str] | None = None,
    ) -> None:
        super().__init__(detail or error_code.value)
        self.error_code = error_code
        self.detail = detail
        self.invalid_cells = list(invalid_cells or [])


class StructureImporter:
    """Handle diagnostic version structure imports from Excel templates."""

    def __init__(self, db: Session) -> None:
        self._db = db

    def import_version_structure(
        self,
        *,
        version_id: int,
        admin_id: int,
        content: bytes,
    ) -> StructureImportSummary:
        version, diagnostic = self._load_version_for_update(version_id)
        batch = self._parse_workbook(content)

        expected_headers = self._expected_outcome_headers(diagnostic.outcome_table_name)
        actual_headers = batch.outcome_headers or self._infer_outcome_headers(batch)
        self._validate_outcome_headers(actual_headers, expected_headers)

        warnings = list(batch.warnings)

        (
            question_count,
            question_map,
            version_question_payloads,
        ) = self._persist_questions(
            version=version,
            diagnostic=diagnostic,
            admin_id=admin_id,
            rows=batch.questions,
        )
        option_count, version_option_payloads = self._persist_options(
            version=version,
            admin_id=admin_id,
            question_map=question_map,
            rows=batch.options,
        )
        (
            outcome_count,
            outcome_warnings,
            version_outcome_payloads,
        ) = self._persist_outcomes(
            version=version,
            diagnostic=diagnostic,
            admin_id=admin_id,
            headers=expected_headers,
            rows=batch.outcomes,
        )

        warnings.extend(outcome_warnings)

        self._replace_version_structure(
            version=version,
            questions=version_question_payloads,
            options=version_option_payloads,
            outcomes=version_outcome_payloads,
        )

        record_diagnostic_version_log(
            self._db,
            version_id=version.id,
            admin_user_id=admin_id,
            action="IMPORT",
            new_value={
                "questions": question_count,
                "options": option_count,
                "outcomes": outcome_count,
                "warnings": warnings,
            },
        )

        return StructureImportSummary(
            version_id=version.id,
            questions_imported=question_count,
            options_imported=option_count,
            outcomes_imported=outcome_count,
            warnings=warnings,
        )

    # ------------------------------------------------------------------#
    # Loading helpers
    # ------------------------------------------------------------------#

    def _load_version_for_update(self, version_id: int) -> tuple[DiagnosticVersion, Diagnostic]:
        stmt = (
            select(DiagnosticVersion, Diagnostic)
            .join(Diagnostic, Diagnostic.id == DiagnosticVersion.diagnostic_id)
            .where(DiagnosticVersion.id == version_id)
            .with_for_update()
        )
        row = self._db.execute(stmt).first()
        if row is None:
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_VERSION_NOT_FOUND,
                detail="Diagnostic version not found",
            )
        version, diagnostic = row
        if version.src_hash:
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_VERSION_FROZEN,
                detail="Diagnostic version is frozen",
            )
        return version, diagnostic

    # ------------------------------------------------------------------#
    # Parsing helpers
    # ------------------------------------------------------------------#

    def _parse_workbook(self, content: bytes) -> StructureImportBatch:
        stream = io.BytesIO(content)
        try:
            workbook = _load_workbook(stream)
        except Exception as exc:  # pragma: no cover - openpyxl provides coarse exceptions
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION,
                detail="Failed to read the uploaded workbook",
            ) from exc

        self._ensure_required_sheets(workbook)
        questions = self._parse_question_sheet(workbook["questions"])
        options = self._parse_option_sheet(workbook["options"])
        outcomes, headers = self._parse_outcome_sheet(workbook["outcomes"])

        return StructureImportBatch(
            questions=questions,
            options=options,
            outcomes=outcomes,
            warnings=[],
            outcome_headers=headers,
        )

    def _ensure_required_sheets(self, workbook: Workbook) -> None:
        missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
        if missing:
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_SHEET_MISSING,
                detail=f"Missing required sheet(s): {', '.join(missing)}",
                invalid_cells=[f"{name}!A1" for name in missing],
            )

    def _parse_question_sheet(self, sheet: Worksheet) -> list[QuestionImportRow]:
        expected = TemplateExporter.QUESTIONS_HEADERS
        headers = self._read_header(sheet)
        self._validate_headers("questions", headers, expected)

        rows: list[QuestionImportRow] = []
        invalid_cells: list[str] = []
        for row_index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue  # header
            q_code, display_text, multi_value, sort_order_value, is_active_value = values[: len(expected)]
            if self._is_row_empty([q_code, display_text, multi_value, sort_order_value, is_active_value]):
                continue
            if not isinstance(q_code, str) or not q_code.strip():
                invalid_cells.append(f"questions!A{row_index}")
                continue
            q_code = q_code.strip()
            if len(q_code) > 64:
                invalid_cells.append(f"questions!A{row_index}")
            multi_flag = self._coerce_bool(multi_value, sheet="questions", column_index=3, row_index=row_index, errors=invalid_cells)
            is_active_flag = self._coerce_bool(
                is_active_value, sheet="questions", column_index=5, row_index=row_index, errors=invalid_cells
            )
            sort_order = self._coerce_int(
                sort_order_value, sheet="questions", column_index=4, row_index=row_index, errors=invalid_cells
            )
            if invalid_cells:
                continue
            rows.append(
                QuestionImportRow(
                    q_code=q_code,
                    display_text=(display_text or "").strip(),
                    multi=multi_flag,
                    sort_order=sort_order,
                    is_active=is_active_flag,
                    row_index=row_index,
                )
            )

        if invalid_cells:
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION,
                detail="questions sheet contains invalid values",
                invalid_cells=invalid_cells,
            )
        return rows

    def _parse_option_sheet(self, sheet: Worksheet) -> list[OptionImportRow]:
        expected = TemplateExporter.OPTIONS_HEADERS
        headers = self._read_header(sheet)
        self._validate_headers("options", headers, expected)

        rows: list[OptionImportRow] = []
        invalid_cells: list[str] = []
        for row_index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue
            q_code, opt_code, display_label, sort_order_value, llm_op_value, is_active_value = values[: len(expected)]
            if self._is_row_empty([q_code, opt_code, display_label, sort_order_value, llm_op_value, is_active_value]):
                continue
            if not isinstance(q_code, str) or not q_code.strip():
                invalid_cells.append(f"options!A{row_index}")
                continue
            if not isinstance(opt_code, str) or not opt_code.strip():
                invalid_cells.append(f"options!B{row_index}")
                continue
            q_code = q_code.strip()
            opt_code = opt_code.strip()
            if len(opt_code) > 64:
                invalid_cells.append(f"options!B{row_index}")
            sort_order = self._coerce_int(
                sort_order_value, sheet="options", column_index=4, row_index=row_index, errors=invalid_cells
            )
            is_active_flag = self._coerce_bool(
                is_active_value, sheet="options", column_index=6, row_index=row_index, errors=invalid_cells
            )
            llm_payload = self._parse_llm_column(
                llm_op_value, sheet="options", column_index=5, row_index=row_index, errors=invalid_cells
            )
            if invalid_cells:
                continue
            rows.append(
                OptionImportRow(
                    q_code=q_code,
                    opt_code=opt_code,
                    display_label=(display_label or "").strip(),
                    sort_order=sort_order,
                    is_active=is_active_flag,
                    llm_op=llm_payload,
                    row_index=row_index,
                )
            )

        if invalid_cells:
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION,
                detail="options sheet contains invalid values",
                invalid_cells=invalid_cells,
            )
        return rows

    def _parse_outcome_sheet(self, sheet: Worksheet) -> tuple[list[OutcomeImportRow], list[str]]:
        headers = self._read_header(sheet)

        outcomes: list[OutcomeImportRow] = []
        for row_index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue
            if self._is_row_empty(values):
                continue
            payload: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                payload[header] = values[idx] if idx < len(values) else None
            outcomes.append(OutcomeImportRow(values=payload, row_index=row_index))

        return outcomes, headers

    def _read_header(self, sheet: Worksheet) -> list[str]:
        first_row = next(sheet.iter_rows(values_only=True), [])
        headers = []
        for cell in first_row:
            if cell is None:
                headers.append("")
            else:
                headers.append(str(cell).strip())
        return headers

    def _validate_headers(self, sheet_name: str, actual: Sequence[str], expected: Sequence[str]) -> None:
        if list(actual[: len(expected)]) == list(expected) and len(actual) == len(expected):
            return

        invalid_cells: list[str] = []
        max_len = max(len(actual), len(expected))
        for index in range(max_len):
            expected_value = expected[index] if index < len(expected) else None
            actual_value = actual[index] if index < len(actual) else None
            if expected_value != actual_value:
                column_letter = _column_letter(index + 1)
                invalid_cells.append(f"{sheet_name}!{column_letter}1")

        raise StructureImportParseError(
            error_code=ErrorCode.DIAGNOSTICS_COL_MISSING,
            detail=f"{sheet_name} sheet headers do not match expected columns",
            invalid_cells=invalid_cells,
        )

    def _is_row_empty(self, row: Iterable[Any]) -> bool:
        return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)

    def _coerce_bool(
        self,
        value: Any,
        *,
        sheet: str,
        column_index: int,
        row_index: int,
        errors: list[str],
    ) -> bool:
        truthy = {"1", "true", "yes", "y", "on"}
        falsy = {"0", "false", "no", "n", "off"}
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        if isinstance(value, (int, float)):
            if int(value) == 1:
                return True
            if int(value) == 0:
                return False
        if isinstance(value, str):
            normalised = value.strip().lower()
            if normalised in truthy:
                return True
            if normalised in falsy:
                return False
        errors.append(f"{sheet}!{_column_letter(column_index)}{row_index}")
        return False

    def _coerce_int(
        self,
        value: Any,
        *,
        sheet: str,
        column_index: int,
        row_index: int,
        errors: list[str],
    ) -> int:
        if value is None or (isinstance(value, str) and not value.strip()):
            return 0
        try:
            return int(value)
        except (TypeError, ValueError):
            errors.append(f"{sheet}!{_column_letter(column_index)}{row_index}")
            return 0

    def _parse_llm_column(
        self,
        value: Any,
        *,
        sheet: str,
        column_index: int,
        row_index: int,
        errors: list[str],
    ) -> dict[str, Any] | None:
        if value is None or value == "":
            return None
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
            except json.JSONDecodeError:
                errors.append(f"{sheet}!{_column_letter(column_index)}{row_index}")
                return None
            if isinstance(parsed, dict):
                return parsed
            errors.append(f"{sheet}!{_column_letter(column_index)}{row_index}")
        return None

    def _infer_outcome_headers(self, batch: StructureImportBatch) -> list[str]:
        if batch.outcomes:
            return list(batch.outcomes[0].values.keys())
        return []

    def _expected_outcome_headers(self, table_name: str) -> list[str]:
        inspector = inspect(self._db.get_bind())
        columns = inspector.get_columns(table_name)
        excluded = {"id", "created_at", "updated_at"}
        names = [col["name"] for col in columns if col["name"] not in excluded]
        ordered = [name for name in names if name not in {"sort_order", "is_active"}]
        if "sort_order" in names and "sort_order" not in ordered:
            ordered.append("sort_order")
        if "is_active" in names and "is_active" not in ordered:
            ordered.append("is_active")
        return ordered

    def _validate_outcome_headers(self, actual: list[str], expected: list[str]) -> None:
        if not expected:
            return
        if actual == expected:
            return
        if len(actual) == len(expected) and set(actual) == set(expected):
            return
        invalid_cells: list[str] = []
        max_len = max(len(actual), len(expected))
        for index in range(max_len):
            exp = expected[index] if index < len(expected) else None
            act = actual[index] if index < len(actual) else None
            if exp != act:
                invalid_cells.append(f"outcomes!{_column_letter(index + 1)}1")
        raise StructureImportParseError(
            error_code=ErrorCode.DIAGNOSTICS_COL_MISSING,
            detail="outcomes sheet headers do not match the outcome table schema",
            invalid_cells=invalid_cells,
        )

    # ------------------------------------------------------------------#
    # Persistence helpers
    # ------------------------------------------------------------------#

    def _persist_questions(
        self,
        *,
        version: DiagnosticVersion,
        diagnostic: Diagnostic,
        admin_id: int,
        rows: Sequence[QuestionImportRow],
    ) -> tuple[int, dict[str, Question], list[dict[str, Any]]]:
        existing = self._db.execute(
            select(Question).where(Question.diagnostic_id == diagnostic.id)
        ).scalars().all()
        questions_by_code = {question.q_code: question for question in existing}

        for row in rows:
            question = questions_by_code.get(row.q_code)
            if question is None:
                question = Question(
                    diagnostic_id=diagnostic.id,
                    q_code=row.q_code,
                    display_text=row.display_text,
                    multi=row.multi,
                    sort_order=row.sort_order,
                    is_active=row.is_active,
                )
                self._db.add(question)
                self._db.flush()
                questions_by_code[row.q_code] = question
            else:
                question.display_text = row.display_text
                question.multi = row.multi
                question.sort_order = row.sort_order
                question.is_active = row.is_active

        self._db.flush()

        version_question_payloads: list[dict[str, Any]] = []
        for row in rows:
            question = questions_by_code[row.q_code]
            version_question_payloads.append(
                {
                    "version_id": version.id,
                    "diagnostic_id": diagnostic.id,
                    "question_id": question.id,
                    "q_code": row.q_code,
                    "display_text": row.display_text,
                    "multi": row.multi,
                    "sort_order": row.sort_order,
                    "is_active": row.is_active,
                    "created_by_admin_id": admin_id,
                }
            )

        return len(rows), questions_by_code, version_question_payloads

    def _persist_options(
        self,
        *,
        version: DiagnosticVersion,
        admin_id: int,
        question_map: dict[str, Question],
        rows: Sequence[OptionImportRow],
    ) -> tuple[int, list[dict[str, Any]]]:
        option_errors: list[str] = []
        options_by_key: dict[tuple[int, str], Option] = {}
        options_by_sort: dict[tuple[int, int], Option] = {}

        question_ids = [question.id for question in question_map.values()]
        if question_ids:
            existing_options = self._db.execute(
                select(Option).where(Option.question_id.in_(question_ids))
            ).scalars().all()
            for option in existing_options:
                options_by_key[(option.question_id, option.opt_code)] = option
                options_by_sort[(option.question_id, option.sort_order)] = option

        for row in rows:
            question = question_map.get(row.q_code)
            if question is None:
                option_errors.append(f"options!A{row.row_index or 2}")
                continue
            key = (question.id, row.opt_code)
            option = options_by_key.get(key)
            if option is None:
                sort_key = (question.id, row.sort_order)
                existing_by_sort = options_by_sort.get(sort_key)
                if existing_by_sort is not None:
                    old_key = (question.id, existing_by_sort.opt_code)
                    if old_key in options_by_key:
                        options_by_key.pop(old_key, None)
                    existing_by_sort.opt_code = row.opt_code
                    existing_by_sort.display_label = row.display_label
                    existing_by_sort.sort_order = row.sort_order
                    existing_by_sort.is_active = row.is_active
                    existing_by_sort.llm_op = row.llm_op
                    options_by_key[key] = existing_by_sort
                    options_by_sort[sort_key] = existing_by_sort
                    continue

                option = Option(
                    question_id=question.id,
                    opt_code=row.opt_code,
                    display_label=row.display_label,
                    sort_order=row.sort_order,
                    is_active=row.is_active,
                    llm_op=row.llm_op,
                )
                self._db.add(option)
                self._db.flush()
                options_by_key[key] = option
                options_by_sort[(question.id, option.sort_order)] = option
            else:
                previous_sort_order = option.sort_order
                option.display_label = row.display_label
                option.sort_order = row.sort_order
                option.is_active = row.is_active
                option.llm_op = row.llm_op
                if previous_sort_order != option.sort_order:
                    options_by_sort.pop((question.id, previous_sort_order), None)
                options_by_sort[(question.id, option.sort_order)] = option

        if option_errors:
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION,
                detail="options sheet refers to unknown question codes",
                invalid_cells=option_errors,
            )

        version_option_payloads: list[dict[str, Any]] = []
        for row in rows:
            question = question_map.get(row.q_code)
            if question is None:
                continue  # validation already handled
            option = options_by_key[(question.id, row.opt_code)]
            version_option_payloads.append(
                {
                    "version_id": version.id,
                    "question_id": question.id,
                    "option_id": option.id,
                    "q_code": row.q_code,
                    "opt_code": row.opt_code,
                    "display_label": row.display_label,
                    "sort_order": row.sort_order,
                    "llm_op": row.llm_op,
                    "is_active": row.is_active,
                    "created_by_admin_id": admin_id,
                }
            )

        return len(version_option_payloads), version_option_payloads

    def _persist_outcomes(
        self,
        *,
        version: DiagnosticVersion,
        diagnostic: Diagnostic,
        admin_id: int,
        headers: Sequence[str],
        rows: Sequence[OutcomeImportRow],
    ) -> tuple[int, list[str], list[dict[str, Any]]]:
        if not headers and rows:
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_COL_MISSING,
                detail="outcomes sheet header is empty",
            )

        binding = resolve_outcome_model(diagnostic.outcome_table_name)
        model = binding.model
        key_columns = binding.key_columns
        if not key_columns:
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_DEP_MISSING,
                detail="Outcome model configuration is missing key columns",
            )

        column_positions = {name: idx for idx, name in enumerate(headers, start=1)}
        invalid_cells: list[str] = []
        missing_key_names: set[str] = set()
        warnings: list[str] = []

        outcome_count = 0
        version_outcome_payloads: list[dict[str, Any]] = []
        for row in rows:
            payload = {name: row.values.get(name) for name in headers}
            missing_keys = [key for key in key_columns if not self._has_value(payload.get(key))]
            if missing_keys:
                missing_key_names.update(missing_keys)
                for key in missing_keys:
                    column = column_positions.get(key, 1)
                    invalid_cells.append(f"outcomes!{_column_letter(column)}{row.row_index or 2}")
                continue

            normalised = self._normalise_outcome_payload(payload)
            filters = [getattr(model, key) == normalised[key] for key in key_columns]
            existing = self._db.execute(select(model).where(*filters)).scalar_one_or_none()
            created = False
            if existing is None:
                instance = model(**normalised)
                self._db.add(instance)
                self._db.flush()
                created = True
            else:
                for key, value in normalised.items():
                    setattr(existing, key, value)
                instance = existing

            sort_order_value = int(normalised.get("sort_order", 0))
            is_active_value = bool(normalised.get("is_active", True))
            payload["sort_order"] = sort_order_value
            payload["is_active"] = 1 if is_active_value else 0

            version_outcome_payloads.append(
                {
                    "version_id": version.id,
                    "outcome_id": getattr(instance, "id"),
                    "sort_order": sort_order_value,
                    "is_active": is_active_value,
                    "outcome_meta_json": payload,
                    "created_by_admin_id": admin_id,
                }
            )
            outcome_count += 1
            if created:
                label = str(normalised.get(binding.default_label_column.key, ""))
                warnings.append(f"Outcome '{label}' を新規登録しました")

        if invalid_cells:
            detail = "outcomes sheet contains rows missing required key values"
            if missing_key_names:
                detail = f"{detail} ({', '.join(sorted(missing_key_names))})"
            raise StructureImportParseError(
                error_code=ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION,
                detail=detail,
                invalid_cells=invalid_cells,
            )

        return outcome_count, warnings, version_outcome_payloads

    def _replace_version_structure(
        self,
        *,
        version: DiagnosticVersion,
        questions: Sequence[dict[str, Any]],
        options: Sequence[dict[str, Any]],
        outcomes: Sequence[dict[str, Any]],
    ) -> None:
        self._db.execute(delete(VersionOutcome).where(VersionOutcome.version_id == version.id))
        self._db.execute(delete(VersionOption).where(VersionOption.version_id == version.id))
        self._db.execute(delete(VersionQuestion).where(VersionQuestion.version_id == version.id))
        self._db.flush()

        version_question_instances: list[VersionQuestion] = []
        for payload in questions:
            instance = VersionQuestion(**payload)
            self._db.add(instance)
            version_question_instances.append(instance)
        self._db.flush()

        question_to_snapshot: dict[int, int] = {
            instance.question_id: instance.id for instance in version_question_instances
        }

        for payload in options:
            question_id = payload.pop("question_id")
            version_question_id = question_to_snapshot.get(question_id)
            if version_question_id is None:  # pragma: no cover - defensive guard
                raise StructureImportParseError(
                    error_code=ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION,
                    detail="options row refers to question missing from version snapshot",
                )
            payload["version_question_id"] = version_question_id
            self._db.add(VersionOption(**payload))
        self._db.flush()

        for payload in outcomes:
            self._db.add(VersionOutcome(**payload))
        self._db.flush()

    def _normalise_outcome_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        normalised: dict[str, Any] = {}
        for key, value in payload.items():
            if key == "is_active":
                normalised[key] = bool(value) if isinstance(value, bool) else str(value).strip() in {"1", "true", "True"}
            elif key == "sort_order":
                try:
                    normalised[key] = int(value)
                except (TypeError, ValueError):
                    normalised[key] = 0
            elif isinstance(value, str):
                normalised[key] = value.strip()
            else:
                normalised[key] = value
        return normalised

    def _has_value(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return bool(value.strip())
        return True


__all__ = [
    "OptionImportRow",
    "OutcomeImportRow",
    "QuestionImportRow",
    "StructureImportBatch",
    "StructureImportParseError",
    "StructureImportSummary",
    "StructureImporter",
]
