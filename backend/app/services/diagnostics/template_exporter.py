from __future__ import annotations

import io
import json
from dataclasses import dataclass
from typing import Any, Sequence
from sqlalchemy import MetaData, Table, inspect, select
from sqlalchemy.orm import Session

from app.core.errors import ErrorCode
from app.core.exceptions import raise_app_error
from app.models.diagnostic import Diagnostic, DiagnosticVersion, Option, Question, VersionOption, VersionOutcome, VersionQuestion

JSON_LIKE = (dict, list)


@dataclass(frozen=True)
class DiagnosticMeta:
    id: int
    code: str
    outcome_table_name: str


@dataclass(frozen=True)
class TemplateResult:
    filename: str
    content: bytes


class TemplateExporter:
    FILE_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    QUESTIONS_HEADERS = ["q_code", "display_text", "multi", "sort_order", "is_active"]
    OPTIONS_HEADERS = ["q_code", "opt_code", "display_label", "sort_order", "llm_op", "is_active"]

    def __init__(self, db: Session) -> None:
        self._db = db

    def build(self, *, version_id: int, diagnostic_id: int | None) -> TemplateResult:
        if version_id > 0:
            diagnostic = self._load_diagnostic_for_version(version_id)
            content = self._build_from_version_data(diagnostic, version_id)
            filename = f"{diagnostic.code}_v{version_id}.xlsx"
            return TemplateResult(filename=filename, content=content)

        if diagnostic_id is None or diagnostic_id <= 0:
            raise_app_error(ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION)

        diagnostic = self._load_diagnostic(diagnostic_id)
        draft_id = self._find_latest_draft_version(diagnostic.id)
        if draft_id is not None:
            content = self._build_from_version_data(diagnostic, draft_id)
        else:
            content = self._build_from_master_data(diagnostic)
        filename = f"{diagnostic.code}_vdraft.xlsx"
        return TemplateResult(filename=filename, content=content)

    def _load_diagnostic_for_version(self, version_id: int) -> DiagnosticMeta:
        stmt = (
            select(
                DiagnosticVersion.id,
                DiagnosticVersion.diagnostic_id,
                Diagnostic.code,
                Diagnostic.outcome_table_name,
            )
            .join(Diagnostic, Diagnostic.id == DiagnosticVersion.diagnostic_id)
            .where(DiagnosticVersion.id == version_id)
        )
        row = self._db.execute(stmt).first()
        if row is None:
            raise_app_error(ErrorCode.DIAGNOSTICS_VERSION_NOT_FOUND)
        _, diagnostic_id, code, outcome_table_name = row
        return DiagnosticMeta(id=diagnostic_id, code=code, outcome_table_name=outcome_table_name)

    def _load_diagnostic(self, diagnostic_id: int) -> DiagnosticMeta:
        row = self._db.execute(
            select(Diagnostic.id, Diagnostic.code, Diagnostic.outcome_table_name).where(
                Diagnostic.id == diagnostic_id
            )
        ).first()
        if row is None:
            raise_app_error(ErrorCode.DIAGNOSTICS_DIAGNOSTIC_NOT_FOUND)
        diag_id, code, outcome_table_name = row
        return DiagnosticMeta(id=diag_id, code=code, outcome_table_name=outcome_table_name)

    def _find_latest_draft_version(self, diagnostic_id: int) -> int | None:
        stmt = (
            select(DiagnosticVersion.id)
            .where(
                DiagnosticVersion.diagnostic_id == diagnostic_id,
                DiagnosticVersion.src_hash.is_(None),
            )
            .order_by(DiagnosticVersion.id.desc())
        )
        return self._db.execute(stmt).scalars().first()

    def _build_from_version_data(self, diagnostic: DiagnosticMeta, version_id: int) -> bytes:
        questions = self._db.execute(
            select(
                VersionQuestion.q_code,
                VersionQuestion.display_text,
                VersionQuestion.multi,
                VersionQuestion.sort_order,
                VersionQuestion.is_active,
            )
            .where(VersionQuestion.version_id == version_id)
            .order_by(VersionQuestion.sort_order, VersionQuestion.id)
        ).all()

        options = self._db.execute(
            select(
                VersionQuestion.q_code,
                VersionOption.opt_code,
                VersionOption.display_label,
                VersionOption.sort_order,
                VersionOption.llm_op,
                VersionOption.is_active,
            )
            .join(
                VersionQuestion,
                VersionQuestion.id == VersionOption.version_question_id,
            )
            .where(VersionOption.version_id == version_id)
            .order_by(
                VersionQuestion.sort_order,
                VersionOption.sort_order,
                VersionOption.id,
            )
        ).all()

        outcomes = self._db.execute(
            select(
                VersionOutcome.outcome_meta_json,
                VersionOutcome.sort_order,
                VersionOutcome.is_active,
            )
            .where(VersionOutcome.version_id == version_id)
            .order_by(VersionOutcome.sort_order, VersionOutcome.outcome_id)
        ).all()

        outcome_headers = self._resolve_outcome_headers(diagnostic.outcome_table_name)
        return self._build_workbook(
            questions=questions,
            options=options,
            outcomes=self._format_version_outcomes(outcomes, outcome_headers),
            outcome_headers=outcome_headers,
        )

    def _build_from_master_data(self, diagnostic: DiagnosticMeta) -> bytes:
        questions = self._db.execute(
            select(
                Question.q_code,
                Question.display_text,
                Question.multi,
                Question.sort_order,
                Question.is_active,
            )
            .where(Question.diagnostic_id == diagnostic.id, Question.is_active.is_(True))
            .order_by(Question.sort_order, Question.id)
        ).all()

        options = self._db.execute(
            select(
                Question.q_code,
                Option.opt_code,
                Option.display_label,
                Option.sort_order,
                Option.llm_op,
                Option.is_active,
            )
            .join(Question, Question.id == Option.question_id)
            .where(
                Question.diagnostic_id == diagnostic.id,
                Question.is_active.is_(True),
                Option.is_active.is_(True),
            )
            .order_by(
                Question.sort_order,
                Question.id,
                Option.sort_order,
                Option.id,
            )
        ).all()

        # Reflect outcome master table for dynamic column handling
        table = self._reflect_outcome_table(diagnostic.outcome_table_name)
        outcome_headers = self._resolve_outcome_headers(diagnostic.outcome_table_name)

        stmt = select(*(table.c[name] for name in outcome_headers if name in table.c))
        if "is_active" in table.c:
            stmt = stmt.where(table.c.is_active.is_(True))
        order_by = []
        if "sort_order" in table.c:
            order_by.append(table.c.sort_order)
        if "id" in table.c:
            order_by.append(table.c.id)
        if order_by:
            stmt = stmt.order_by(*order_by)
        outcomes = self._db.execute(stmt).all()

        formatted_outcomes = [
            [
                self._normalise_outcome_cell(name, row._mapping.get(name))
                for name in outcome_headers
            ]
            for row in outcomes
        ]

        return self._build_workbook(
            questions=questions,
            options=options,
            outcomes=formatted_outcomes,
            outcome_headers=outcome_headers,
        )

    def _format_version_outcomes(
        self,
        outcomes: Sequence[tuple[dict[str, Any] | None, int, bool]],
        headers: list[str],
    ) -> list[list[Any]]:
        formatted: list[list[Any]] = []
        for meta, sort_order, is_active in outcomes:
            payload = dict(meta or {})
            payload["sort_order"] = sort_order
            payload["is_active"] = 1 if is_active else 0
            row = [self._normalise_outcome_cell(name, payload.get(name)) for name in headers]
            formatted.append(row)
        return formatted

    def _build_workbook(
        self,
        *,
        questions: Sequence[Sequence[Any]],
        options: Sequence[Sequence[Any]],
        outcomes: Sequence[Sequence[Any]],
        outcome_headers: list[str],
    ) -> bytes:
        from openpyxl import Workbook

        wb = Workbook()
        ws_questions = wb.active
        ws_questions.title = "questions"
        ws_questions.append(self.QUESTIONS_HEADERS)
        for q_code, text, multi, sort_order, is_active in questions:
            ws_questions.append(
                [
                    q_code,
                    text,
                    self._bool_to_flag(multi),
                    sort_order,
                    self._bool_to_flag(is_active),
                ]
            )

        ws_options = wb.create_sheet("options")
        ws_options.append(self.OPTIONS_HEADERS)
        for q_code, opt_code, label, sort_order, llm_op, is_active in options:
            ws_options.append(
                [
                    q_code,
                    opt_code,
                    label,
                    sort_order,
                    self._dump_json(llm_op),
                    self._bool_to_flag(is_active),
                ]
            )

        ws_outcomes = wb.create_sheet("outcomes")
        ws_outcomes.append(outcome_headers)
        for row in outcomes:
            ws_outcomes.append(row)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _dump_json(self, value: Any) -> str:
        if value is None or value == "":
            return ""
        if isinstance(value, JSON_LIKE):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        return str(value)

    def _bool_to_flag(self, value: Any) -> int:
        return 1 if bool(value) else 0

    def _resolve_outcome_headers(self, table_name: str) -> list[str]:
        inspector = inspect(self._db.get_bind())
        columns = inspector.get_columns(table_name)
        excluded = {"id", "created_at", "updated_at"}
        names: list[str] = [col["name"] for col in columns if col["name"] not in excluded]
        ordered: list[str] = [name for name in names if name not in {"sort_order", "is_active"}]
        if "sort_order" not in ordered and "sort_order" in names:
            ordered.append("sort_order")
        if "is_active" not in ordered and "is_active" in names:
            ordered.append("is_active")
        return ordered

    def _reflect_outcome_table(self, table_name: str) -> Table:
        metadata = MetaData()
        return Table(table_name, metadata, autoload_with=self._db.get_bind())

    def _normalise_outcome_cell(self, name: str, value: Any) -> Any:
        if value is None:
            return None
        if name in {"is_active", "multi"}:
            return self._bool_to_flag(value)
        return value
