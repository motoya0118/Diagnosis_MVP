from __future__ import annotations

import io
import json
import os
from collections.abc import Iterator

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, delete, event, func, select
from sqlalchemy.orm import Session, sessionmaker

from app.core.errors import ErrorCode
from app.core.security import create_access_token
from app.deps import admin as admin_deps
from app.main import app
from app.models.diagnostic import (
    Diagnostic,
    DiagnosticVersion,
    DiagnosticVersionAuditLog,
    Question,
    Option,
    VersionOption,
    VersionOutcome,
    VersionQuestion,
)
from app.models.mst_ai_job import MstAiJob
from tests.factories import (
    AdminUserFactory,
    DiagnosticFactory,
    DiagnosticVersionFactory,
    set_factory_session,
)


def _get_database_url() -> str:
    url = os.environ.get("TEST_DATABASE_URL") or os.environ.get("DATABASE_URL")
    assert url, "DATABASE_URL or TEST_DATABASE_URL must be set for tests"
    return url


@pytest.fixture
def db_session() -> Iterator[Session]:
    engine = create_engine(_get_database_url(), future=True)
    connection = engine.connect()
    transaction = connection.begin()

    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=connection, future=True)
    session = TestingSessionLocal()
    session.begin_nested()

    @event.listens_for(session, "after_transaction_end")
    def restart_savepoint(sess, trans):  # pragma: no cover - fixture wiring
        if trans.nested and not trans._parent.nested:
            sess.begin_nested()

    for model in [
        VersionOutcome,
        VersionOption,
        VersionQuestion,
        DiagnosticVersionAuditLog,
        DiagnosticVersion,
        Option,
        Question,
        MstAiJob,
        Diagnostic,
    ]:
        session.execute(delete(model))
    session.flush()
    set_factory_session(session)

    try:
        yield session
    finally:
        session.rollback()
        session.close()
        transaction.rollback()
        connection.close()
        engine.dispose()
        set_factory_session(None)


@pytest.fixture
def client(db_session: Session) -> Iterator[TestClient]:
    def override_get_db() -> Iterator[Session]:
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[admin_deps.get_db] = override_get_db
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(admin_deps.get_db, None)


def _auth_header(admin_id: int, role: str = "admin", user_id: str | None = None) -> dict[str, str]:
    token = create_access_token(
        str(admin_id),
        extra={"role": role, "user_id": user_id or f"admin{admin_id:03d}"},
        expires_delta_minutes=15,
    )
    return {"Authorization": f"Bearer {token}"}


def _load_workbook(content: bytes) -> openpyxl.Workbook:
    with io.BytesIO(content) as stream:
        return openpyxl.load_workbook(stream, data_only=True)


def _sheet_rows(wb: openpyxl.Workbook, name: str) -> tuple[list[str], list[list]]:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    body = [list(row) for row in rows[1:]]
    return headers, body


def _make_question(
    db: Session,
    diagnostic_id: int,
    code: str,
    text: str,
    *,
    multi: bool,
    sort_order: int,
    is_active: bool,
) -> Question:
    question = Question(
        diagnostic_id=diagnostic_id,
        q_code=code,
        display_text=text,
        multi=multi,
        sort_order=sort_order,
        is_active=is_active,
    )
    db.add(question)
    db.flush()
    return question


def _make_option(
    db: Session,
    question: Question,
    *,
    code: str,
    label: str,
    sort_order: int,
    is_active: bool = True,
    llm_op: dict | None = None,
) -> Option:
    option = Option(
        question_id=question.id,
        opt_code=code,
        display_label=label,
        sort_order=sort_order,
        is_active=is_active,
        llm_op=llm_op,
    )
    db.add(option)
    db.flush()
    return option


EXPECTED_OUTCOME_HEADERS = [
    "name",
    "category",
    "role_summary",
    "main_role",
    "collaboration_style",
    "strength_areas",
    "description",
    "avg_salary_jpy",
    "target_phase",
    "core_skills",
    "deliverables",
    "pathway_detail",
    "ai_tools",
    "advice",
    "sort_order",
    "is_active",
]


def _sample_outcome_payload(**overrides: object) -> dict[str, object]:
    base: dict[str, object] = {
        "name": "AI Strategist",
        "category": "戦略・経営系",
        "role_summary": "Leads AI adoption",
        "main_role": "企業全体のAI推進を統括する",
        "collaboration_style": "経営層・事業部門と連携",
        "strength_areas": "戦略立案, リーダーシップ",
        "description": "Guides organizations",
        "avg_salary_jpy": "9000000",
        "target_phase": "拡大型",
        "core_skills": "Strategy, AI",
        "deliverables": "AIロードマップ",
        "pathway_detail": "Start in data science",
        "ai_tools": "ChatGPT",
        "advice": "Keep learning",
        "sort_order": 10,
        "is_active": 1,
    }
    base.update(overrides)
    return base


def _build_ai_job(**overrides: object) -> MstAiJob:
    defaults: dict[str, object] = {
        "name": "Sample Role",
        "category": "戦略・経営系",
        "role_summary": "Summary",
        "main_role": "Main responsibilities",
        "collaboration_style": "Stakeholder engagement",
        "strength_areas": "Strength areas",
        "description": "Detailed description",
        "avg_salary_jpy": "8000000",
        "target_phase": "拡大型",
        "core_skills": "Core skills",
        "deliverables": "Key deliverables",
        "pathway_detail": "Career pathway",
        "ai_tools": "ChatGPT",
        "advice": "Advice message",
        "is_active": True,
        "sort_order": 1,
    }
    defaults.update(overrides)
    return MstAiJob(**defaults)


def test_get_template_for_version_returns_excel(client: TestClient, db_session: Session) -> None:
    admin = AdminUserFactory(is_active=True)
    diagnostic = DiagnosticFactory(code="career", outcome_table_name="mst_ai_jobs")
    version = DiagnosticVersionFactory(
        diagnostic=diagnostic,
        created_by_admin=admin,
        updated_by_admin=admin,
    )

    q1 = _make_question(
        db_session,
        diagnostic.id,
        code="Q001",
        text="What is your goal?",
        multi=False,
        sort_order=1,
        is_active=True,
    )
    option = _make_option(
        db_session,
        q1,
        code="A",
        label="Explore AI careers",
        sort_order=1,
        llm_op={"weight": 0.4},
    )

    vq = VersionQuestion(
        version_id=version.id,
        diagnostic_id=diagnostic.id,
        question_id=q1.id,
        q_code=q1.q_code,
        display_text=q1.display_text,
        multi=q1.multi,
        sort_order=1,
        is_active=True,
        created_by_admin_id=admin.id,
    )
    db_session.add(vq)
    db_session.flush()

    vo = VersionOption(
        version_id=version.id,
        version_question_id=vq.id,
        option_id=option.id,
        q_code=q1.q_code,
        opt_code=option.opt_code,
        display_label=option.display_label,
        llm_op=option.llm_op,
        sort_order=1,
        is_active=True,
        created_by_admin_id=admin.id,
    )
    db_session.add(vo)

    outcome_payload = _sample_outcome_payload()
    vo_meta = VersionOutcome(
        version_id=version.id,
        outcome_id=101,
        outcome_meta_json=outcome_payload,
        sort_order=10,
        created_by_admin_id=admin.id,
    )
    db_session.add(vo_meta)
    db_session.commit()

    headers = _auth_header(admin.id, user_id=admin.user_id)
    response = client.get(f"/admin/diagnostics/versions/{version.id}/template", headers=headers)

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["content-disposition"].endswith(f'{diagnostic.code}_v{version.id}.xlsx"')

    wb = _load_workbook(response.content)

    question_headers, question_rows = _sheet_rows(wb, "questions")
    assert question_headers == ["q_code", "display_text", "multi", "sort_order", "is_active"]
    assert question_rows == [["Q001", "What is your goal?", 0, 1, 1]]

    option_headers, option_rows = _sheet_rows(wb, "options")
    assert option_headers == ["q_code", "opt_code", "display_label", "sort_order", "llm_op", "is_active"]
    assert option_rows == [["Q001", "A", "Explore AI careers", 1, json.dumps({"weight": 0.4}, ensure_ascii=False, sort_keys=True), 1]]

    outcome_headers, outcome_rows = _sheet_rows(wb, "outcomes")
    assert outcome_headers == EXPECTED_OUTCOME_HEADERS
    assert outcome_rows == [[outcome_payload[key] for key in EXPECTED_OUTCOME_HEADERS]]


def test_get_template_for_initial_version_uses_master_data(client: TestClient, db_session: Session) -> None:
    admin = AdminUserFactory(is_active=True)
    diagnostic = DiagnosticFactory(code="talent", outcome_table_name="mst_ai_jobs")

    q_active = _make_question(
        db_session,
        diagnostic.id,
        code="Q100",
        text="Primary focus?",
        multi=True,
        sort_order=1,
        is_active=True,
    )
    _make_question(
        db_session,
        diagnostic.id,
        code="Q101",
        text="Inactive question",
        multi=False,
        sort_order=2,
        is_active=False,
    )
    _make_option(
        db_session,
        q_active,
        code="A1",
        label="AI Strategy",
        sort_order=1,
        is_active=True,
        llm_op={"topic": "strategy"},
    )
    _make_option(
        db_session,
        q_active,
        code="A2",
        label="Inactive option",
        sort_order=2,
        is_active=False,
    )

    active_job = _build_ai_job(
        name="AI Consultant",
        role_summary="Advise clients",
        main_role="AI戦略の立案と推進",
        collaboration_style="経営層との合意形成",
        strength_areas="コンサルティング, 戦略立案",
        description="Consulting on AI initiatives",
        avg_salary_jpy="12000000",
        target_phase="拡大型",
        core_skills="Consulting",
        deliverables="戦略提案書",
        pathway_detail="Grow from data analyst",
        ai_tools="ChatGPT",
        advice="Stay curious",
        is_active=True,
        sort_order=5,
    )
    inactive_job = _build_ai_job(
        name="Legacy Role",
        role_summary="Deprecated",
        main_role="Legacy tasks",
        collaboration_style="n/a",
        strength_areas="n/a",
        description="No longer used",
        avg_salary_jpy=None,
        target_phase=None,
        core_skills=None,
        deliverables=None,
        pathway_detail=None,
        ai_tools=None,
        advice=None,
        is_active=False,
        sort_order=99,
    )
    db_session.add_all([active_job, inactive_job])
    db_session.commit()

    headers = _auth_header(admin.id, user_id=admin.user_id)
    response = client.get(
        f"/admin/diagnostics/versions/0/template?diagnostic_id={diagnostic.id}",
        headers=headers,
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-disposition"].endswith(f'{diagnostic.code}_vdraft.xlsx"')

    wb = _load_workbook(response.content)

    question_headers, question_rows = _sheet_rows(wb, "questions")
    assert question_headers == ["q_code", "display_text", "multi", "sort_order", "is_active"]
    assert question_rows == [["Q100", "Primary focus?", 1, 1, 1]]

    option_headers, option_rows = _sheet_rows(wb, "options")
    assert option_headers == ["q_code", "opt_code", "display_label", "sort_order", "llm_op", "is_active"]
    assert option_rows == [[
        "Q100",
        "A1",
        "AI Strategy",
        1,
        json.dumps({"topic": "strategy"}, ensure_ascii=False, sort_keys=True),
        1,
    ]]

    outcome_headers, outcome_rows = _sheet_rows(wb, "outcomes")
    assert outcome_headers == EXPECTED_OUTCOME_HEADERS
    expected_row = [
        active_job.name,
        active_job.category,
        active_job.role_summary,
        active_job.main_role,
        active_job.collaboration_style,
        active_job.strength_areas,
        active_job.description,
        active_job.avg_salary_jpy,
        active_job.target_phase,
        active_job.core_skills,
        active_job.deliverables,
        active_job.pathway_detail,
        active_job.ai_tools,
        active_job.advice,
        active_job.sort_order,
        1,
    ]
    assert outcome_rows == [expected_row]


def test_get_template_requires_diagnostic_id_for_draft(client: TestClient, db_session: Session) -> None:
    admin = AdminUserFactory(is_active=True)
    DiagnosticFactory()

    headers = _auth_header(admin.id, user_id=admin.user_id)
    response = client.get("/admin/diagnostics/versions/0/template", headers=headers)

    assert response.status_code == ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION.http_status
    payload = response.json()
    assert payload["error"]["code"] == ErrorCode.DIAGNOSTICS_IMPORT_VALIDATION.value


def test_get_template_rejects_unknown_version(client: TestClient, db_session: Session) -> None:
    admin = AdminUserFactory(is_active=True)
    DiagnosticFactory()

    headers = _auth_header(admin.id, user_id=admin.user_id)
    response = client.get("/admin/diagnostics/versions/999999/template", headers=headers)

    assert response.status_code == ErrorCode.DIAGNOSTICS_VERSION_NOT_FOUND.http_status
    payload = response.json()
    assert payload["error"]["code"] == ErrorCode.DIAGNOSTICS_VERSION_NOT_FOUND.value


def test_get_template_rejects_unknown_diagnostic_for_draft(client: TestClient, db_session: Session) -> None:
    admin = AdminUserFactory(is_active=True)
    DiagnosticFactory()

    max_id = db_session.execute(select(func.max(Diagnostic.id))).scalar_one()
    headers = _auth_header(admin.id, user_id=admin.user_id)
    response = client.get(
        f"/admin/diagnostics/versions/0/template?diagnostic_id={max_id + 1}",
        headers=headers,
    )

    assert response.status_code == ErrorCode.DIAGNOSTICS_DIAGNOSTIC_NOT_FOUND.http_status
    payload = response.json()
    assert payload["error"]["code"] == ErrorCode.DIAGNOSTICS_DIAGNOSTIC_NOT_FOUND.value
